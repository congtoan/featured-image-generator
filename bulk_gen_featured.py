# -*- coding: utf-8 -*-
"""
Created on Wed Jul  3 19:04:17 2024

@author: Admin
"""
import openpyxl
from slugify import slugify
from gen_featured_img import generate_featured_image

with open("input_title.txt", "r", encoding="utf-8") as file:
    titles = file.read().splitlines()

output_directory = "output/"
excel_filename = "output.xlsx"

# Initialize Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Featured Images"

# Write headers
worksheet.cell(row=1, column=1, value="Title")
worksheet.cell(row=1, column=2, value="Image Filename")

# Generate featured images and populate Excel sheet
for index, title in enumerate(titles, start=2):  # Start from row 2 for data
    print(title)
    slugified_title = slugify(title)
    image_filename = f"{slugified_title}.jpg"
    output_path = f"{output_directory}{image_filename}"

    # Generate featured image
    generate_featured_image(title, output_path)

    # Write data to Excel
    worksheet.cell(row=index, column=1, value=title)
    worksheet.cell(row=index, column=2, value=image_filename)

# Save Excel workbook
workbook.save(excel_filename)

print("All is done")